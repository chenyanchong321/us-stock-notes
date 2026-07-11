#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每周自动校准「市场规模」卡片的绝对锚点（config/marketscale.json）。

背景：市场规模卡走「真实市值锚点 × 大盘指数浮动」架构（见 build_market_scale）。
指数浮动是全自动的，但绝对锚点过去靠人工季度重校——本脚本消灭这最后一块人工。

各市场官方数据源（2026-07-12 调研定稿，详见 市场规模数据源_求助文档.md）：
- 港股：香港金管局 HKMA 开放 API（免注册），月报含「全市场总市值 + 当月末恒指」成对数据
        → 官方每月亲手给一次配对锚，质量最高。
- 日本：JPX 官方 Excel historical-jika.xlsx（固定网址，每月1日更新，月末总市值）
        → 配对锚 = 该月末的 N225 收盘（Yahoo 历史）+ 当日 JPYUSD。
- A股：上交所 commonQuery（股票总貌）+ 深交所 ShowReport（市场总貌），总市值单位亿元人民币
        → 官方口径日更；GH Actions 美国 runner 若连不上中国官网会自动跳过（保留旧锚）。
- 韩国：KRX 数据门户 getJsonData.cmd 全指数快照，코스피/코스닥 行的上市市值 + KOSPI 收盘
        → 美国可达性待验证，失败自动跳过。
- 美股：无免费自动的"绝对总市值"源（companiesmarketcap 反爬、FRED 已停 Wilshire）。
        本脚本只做漂移优化：探测 Yahoo 是否有 Wilshire 5000 全市场指数（^FTW5000/^W5000），
        有则把美股的浮动指数从 ^GSPC 迁移为全市场指数（绝对锚不变，漂移大幅变小）。
        绝对锚的定期核对交给桌面端 AI 定时任务（月检，偏差>3%才改）。

原则（沿袭仓库惯例）：
- 任何源失败/解析异常/数值出界 → 保留旧锚点 + 记日志，绝不静默造数。
- 每次运行把逐市场诊断写进 data/anchors_log.json（成功值、失败原因、单位推断过程），
  远程排障只需看这个文件，不用翻 Actions 日志。
- 幂等：官方月报没出新数时，config 内容不变 → 工作流不产生提交。
"""
import json, io, calendar, datetime, urllib.request, urllib.parse, pathlib, sys, re

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from update_quotes import fetch_series, UA   # 复用主流水线的 Yahoo 抓取（含重试）

ROOT = pathlib.Path(__file__).resolve().parent.parent
CFG_PATH = ROOT / "config/marketscale.json"
LOG_PATH = ROOT / "data/anchors_log.json"

# 各市场合理区间（万亿美元）：解析结果出界=解析错了，拒绝入库
SANITY = {"美股": (40, 150), "A股": (8, 30), "日本": (4, 15), "港股": (3, 12), "韩国": (1, 6)}


import http.cookiejar
_JAR = http.cookiejar.CookieJar()
_OPENER = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_JAR))

def http_get(url, headers=None, data=None, timeout=30, binary=False):
    req = urllib.request.Request(url, headers={**UA, **(headers or {})},
                                 data=data.encode() if isinstance(data, str) else data)
    with _OPENER.open(req, timeout=timeout) as r:
        raw = r.read()
    return raw if binary else raw.decode("utf-8", "replace")


def http_json(url, **kw):
    return json.loads(http_get(url, **kw))


def yahoo_last(sym):
    """最新 (date, close)。"""
    s = fetch_series(sym, "1mo", "1d")
    if not s:
        raise RuntimeError(f"Yahoo {sym} 无数据")
    t, c = s[-1]
    return datetime.datetime.fromtimestamp(t, datetime.timezone.utc).date(), float(c)


def yahoo_close_on(sym, day, rng="6mo"):
    """day（date）当天或之前最近交易日的收盘。"""
    s = fetch_series(sym, rng, "1d")
    if not s:
        raise RuntimeError(f"Yahoo {sym} 无数据")
    best = None
    for t, c in s:
        d = datetime.datetime.fromtimestamp(t, datetime.timezone.utc).date()
        if d <= day:
            best = (d, float(c))
    if not best:
        raise RuntimeError(f"Yahoo {sym} 在 {day} 前无交易日数据")
    return best


def month_end(ym):
    """'2026-06' → date(2026,6,30)"""
    y, m = int(ym[:4]), int(ym[5:7])
    return datetime.date(y, m, calendar.monthrange(y, m)[1])


def _num(x):
    """'43,324,536.7' → float；解析不了返回 None。"""
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "").replace("，", "").strip())
    except ValueError:
        return None


def infer_unit_to_usd_t(value, fx_usd, candidates, lo, hi, log):
    """单位推断：value×候选倍数×汇率，落在 [lo,hi] 万亿美元的那个才是对的单位。
    candidates = [(倍数, 单位名), ...]；推断过程写进 log 便于远程排障。"""
    for mult, name in candidates:
        t = value * mult * fx_usd / 1e12
        log.append(f"按{name}解释 → {t:.2f}万亿美元")
        if lo <= t <= hi:
            log.append(f"✓ 单位={name}")
            return t
    raise RuntimeError(f"单位推断失败：{value} 无一候选落入 [{lo},{hi}] 万亿美元")


# ---------------- 各市场校锚 ----------------

def anchor_hk(diag):
    j = http_json("https://api.hkma.gov.hk/public/market-data-and-statistics/"
                  "monthly-statistical-bulletin/financial/capital-market-statistics"
                  "?pagesize=1&sortby=end_of_month&sortorder=desc")
    rec = j["result"]["records"][0]
    cap_mn, hsi = _num(rec["eq_mkt_ttl_stock_cap"]), _num(rec["eq_mkt_hs_index"])
    if not cap_mn or not hsi:
        raise RuntimeError(f"HKMA 字段缺失: {rec}")
    _, fx = yahoo_last("HKDUSD=X")                     # 港币联系汇率，波动<1%，用最新价即可
    cap_t = cap_mn * 1e6 * fx / 1e12
    diag.append(f"HKMA {rec['end_of_month']}: {cap_mn:,.0f} 百万港元 × {fx:.4f} = {cap_t:.2f}T")
    return {"cap_usd_t": round(cap_t, 2), "index_anchor": hsi,
            "anchor_date": str(month_end(rec["end_of_month"])),
            "anchor_source": "HKMA月报(市值+恒指官方配对)"}


def anchor_cn(diag):
    # 路线一（首选）：读 ECS 哨站写回的 data/cn_mcap.json。
    # 深交所对海外IP软拒（连通但空表），GH Actions 直抓不可靠；杭州 ECS 抓沪深两所零障碍，
    # 每周一/四 18:07 经 GitHub API 写回该文件（安装说明见 交接给ECS_A股市值哨站.md）。
    try:
        cn = json.loads((ROOT / "data/cn_mcap.json").read_text(encoding="utf-8"))
        d0 = datetime.date.fromisoformat(cn["date"])
        age = (datetime.date.today() - d0).days
        if age <= 10 and cn.get("total_yi"):
            _, fx = yahoo_last("CNYUSD=X")
            cap_t = float(cn["total_yi"]) * 1e8 * fx / 1e12
            dd, idx = yahoo_close_on("000300.SS", d0)
            diag.append(f"ECS哨站 {cn['date']}: 沪+深 {cn['total_yi']:,.0f} 亿元 × {fx:.4f} = {cap_t:.2f}T；沪深300@{dd}={idx}")
            return {"cap_usd_t": round(cap_t, 2), "index_anchor": idx, "anchor_date": str(dd),
                    "anchor_source": "ECS哨站(沪深交易所官方)"}
        diag.append(f"ECS哨站数据过期（{cn.get('date')}，{age}天前），转直抓")
    except FileNotFoundError:
        diag.append("ECS哨站文件不存在（尚未安装），转直抓")
    except Exception as e:
        diag.append(f"ECS哨站数据读取失败: {str(e)[:80]}，转直抓")

    # 路线二（兜底）：GH Actions 直抓（深交所大概率空表，届时保留旧锚）
    # 上交所：股票总貌（TOTAL_VALUE=总市值，亿元）。
    # 中国官网对海外 IP 常见间歇性拒连（Connection reset）：https/http 双协议 × 3 次退避重试。
    q = urllib.parse.urlencode({"sqlId": "COMMON_SSE_SJ_GPSJ_GPSJZM_TJSJ_L",
                                "PRODUCT_NAME": "股票,主板,科创板", "type": "inParams"})
    j, last_err = None, None
    for attempt in range(3):
        for scheme in ("https", "http"):
            try:
                j = http_json(f"{scheme}://query.sse.com.cn/commonQuery.do?{q}",
                              headers={"Referer": f"{scheme}://www.sse.com.cn/"}, timeout=20)
                break
            except Exception as e:
                last_err = e
                diag.append(f"上交所 {scheme} 第{attempt+1}次失败: {str(e)[:80]}")
        if j:
            break
        import time as _t; _t.sleep(3 * (attempt + 1))
    if not j:
        raise RuntimeError(f"上交所三轮重试均失败（海外IP大概率被拒，考虑ECS路线）: {last_err}")
    rows = j.get("result") or []
    row = next((r for r in rows if str(r.get("PRODUCT_NAME")) == "股票"), None)
    if not row:
        raise RuntimeError(f"上交所返回无「股票」行: {str(rows)[:200]}")
    sse_yi = next((v for v in (_num(row.get(k)) for k in
                   ("TOTAL_VALUE", "MARKET_VALUE", "TOTAL_MARKET_VALUE")) if v), None)
    if not sse_yi:   # 字段名变了就在所有含 VALUE 的键里找量级合理的（沪市总市值 40万亿~90万亿元 → 4e5~9e5 亿元）
        cand = {k: _num(v) for k, v in row.items() if "VALUE" in k.upper()}
        sse_yi = next((v for v in cand.values() if v and 3e5 < v < 1.2e6), None)
        diag.append(f"上交所字段兜底匹配: {cand}")
    if not sse_yi:
        raise RuntimeError(f"上交所总市值字段解析失败: {row}")

    # 深交所：市场总貌（找“股票”行的“总市值”列，亿元）。
    # 必须带 txtQueryDate=交易日，否则周末/盘前返回空表（recordcount=0）；逐日回退找最近交易日。
    # 同样对海外IP间歇拒连 → 每个日期最多重试2次。
    j2 = None
    for back in range(0, 7):
        qd = (datetime.date.today() - datetime.timedelta(days=back)).isoformat()
        for attempt in range(2):
            try:
                cand = http_json("https://www.szse.cn/api/report/ShowReport/data"
                                 f"?SHOWTYPE=JSON&CATALOGID=1803_sczm&TABKEY=tab1&txtQueryDate={qd}",
                                 headers={"Referer": "https://www.szse.cn/"}, timeout=20)
                if any((t.get("metadata") or {}).get("recordcount", 0) > 0
                       for t in cand if isinstance(t, dict)):
                    j2 = cand
                    diag.append(f"深交所取到 {qd} 数据")
                else:
                    diag.append(f"深交所 {qd} 空表（非交易日），回退前一日")
                break
            except Exception as e:
                last_err = e
                diag.append(f"深交所 {qd} 第{attempt+1}次失败: {str(e)[:80]}")
                import time as _t; _t.sleep(3)
        if j2:
            break
    if j2 is None:
        raise RuntimeError(f"深交所7天内无有效数据（海外IP被拒或接口变更）: {last_err}")
    # 结构无关的防御式解析：递归找到含「股票」的字典行，取行内量级合理的最大数
    # （总市值≥流通市值，深市总市值 20万亿~60万亿元 → 1.5e5~7e5 亿元；成交额/数量不在此量级）
    def _iter_dicts(o):
        if isinstance(o, dict):
            yield o
            for v in o.values():
                yield from _iter_dicts(v)
        elif isinstance(o, list):
            for v in o:
                yield from _iter_dicts(v)
    szse_yi = None
    for d0 in _iter_dicts(j2):
        vals = list(d0.values())
        if any(isinstance(v, str) and v.strip() == "股票" for v in vals):
            diag.append(f"深交所股票行: {json.dumps(d0, ensure_ascii=False)[:220]}")
            cand = [x for x in (_num(v) for v in vals) if x and 1.5e5 < x < 7e5]
            if cand:
                szse_yi = max(cand)
                break
    if not szse_yi:
        raise RuntimeError(f"深交所总市值解析失败，返回结构: {str(j2)[:300]}")

    _, fx = yahoo_last("CNYUSD=X")
    cap_t = (sse_yi + szse_yi) * 1e8 * fx / 1e12
    d, idx = yahoo_last("000300.SS")
    diag.append(f"沪 {sse_yi:,.0f} + 深 {szse_yi:,.0f} 亿元 × {fx:.4f} = {cap_t:.2f}T；沪深300@{d}={idx}")
    return {"cap_usd_t": round(cap_t, 2), "index_anchor": idx, "anchor_date": str(d),
            "anchor_source": "上交所+深交所官方总貌"}


def anchor_jp(diag):
    raw = http_get("https://www.jpx.co.jp/english/markets/statistics-equities/misc/"
                   "tvdivq0000001w3y-att/historical-jika.xlsx", binary=True)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    diag.append(f"JPX xlsx sheets: {wb.sheetnames}")
    best = None   # (date, value)
    for ws in wb.worksheets:
        header_total_col = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
            for cell in row:
                if isinstance(cell.value, str) and re.search(r"total|合計|総合計", cell.value, re.I):
                    header_total_col = cell.column
        # 从底往上找最近一行「日期 + Total 数值」
        for row in reversed(list(ws.iter_rows(min_row=1, max_row=ws.max_row))):
            vals = [c.value for c in row]
            day = None
            for v in vals:
                if isinstance(v, datetime.datetime):
                    day = v.date(); break
                if isinstance(v, str):
                    m = re.match(r"(\d{4})[./-](\d{1,2})", v.strip())
                    if m:
                        day = month_end(f"{m.group(1)}-{int(m.group(2)):02d}"); break
            if not day:
                continue
            if header_total_col:
                tv = _num(row[header_total_col - 1].value if header_total_col - 1 < len(row) else None)
            else:
                nums = [_num(v) for v in vals]
                nums = [n for n in nums if n and n > 1e4]
                tv = max(nums) if nums else None   # 无 Total 表头时取行内最大数（总计≥任何分市场）
            if tv:
                if not best or day > best[0]:
                    best = (day, tv, ws.title)
                break   # 该 sheet 最近一行已找到
    if not best:
        raise RuntimeError("JPX xlsx 未解析出「日期+总市值」行")
    day, tv, sheet = best
    diag.append(f"JPX {sheet} 最近行: {day} 原始值 {tv:,.0f}")
    _, fxd = yahoo_close_on("JPYUSD=X", day)
    lo, hi = SANITY["日本"]
    cap_t = infer_unit_to_usd_t(tv, fxd, [(1e6, "百万円"), (1e8, "億円"), (1e12, "兆円"), (1e3, "千円")],
                                lo, hi, diag)
    nd, n225 = yahoo_close_on("^N225", day)
    return {"cap_usd_t": round(cap_t, 2), "index_anchor": n225, "anchor_date": str(nd),
            "anchor_source": "JPX官方月末市值Excel"}


def anchor_kr(diag):
    # KRX 数据门户全指数快照：코스피/코스닥 行含上市市值与收盘。逐日回退找最近交易日。
    # 配方来自 pykrx 源码（website/comm/webio.py + krxio.py）：
    # https + Referer=outerLoader + X-Requested-With，参数只需 bld/trdDd/idxIndMidclssCd
    url = "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
    hdr = {"Referer": "https://data.krx.co.kr/contents/MDC/MDI/outerLoader/index.cmd",
           "X-Requested-With": "XMLHttpRequest",
           "Content-Type": "application/x-www-form-urlencoded"}
    # Cookie 预热：拿会话
    try:
        http_get("https://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd?menuId=MDC0201", timeout=15)
        diag.append("KRX 预热成功（已取会话Cookie）")
    except Exception as e:
        diag.append(f"KRX 预热失败: {str(e)[:80]}")
    today = datetime.date.today()
    for back in range(1, 8):
        trd = (today - datetime.timedelta(days=back)).strftime("%Y%m%d")
        got = {}
        for mid, label in (("02", "코스피"), ("03", "코스닥")):
            body = urllib.parse.urlencode({"bld": "dbms/MDC/STAT/standard/MDCSTAT00101",
                                           "idxIndMidclssCd": mid, "trdDd": trd})
            j = http_json(url, headers=hdr, data=body)
            rows = j.get("output") or j.get("OutBlock_1") or []
            row = next((r for r in rows if str(r.get("IDX_NM", "")).strip() == label), None)
            if row:
                got[label] = row
        if len(got) == 2:
            mk = _num(got["코스피"].get("MKTCAP")); mq = _num(got["코스닥"].get("MKTCAP"))
            kospi = _num(got["코스피"].get("CLSPRC_IDX"))
            if mk and mq and kospi:
                _, fx = yahoo_last("KRWUSD=X")
                total = mk + mq
                diag.append(f"KRX {trd}: 코스피 {mk:,.0f} + 코스닥 {mq:,.0f}（原始单位待推断）KOSPI={kospi}")
                lo, hi = SANITY["韩国"]
                cap_t = infer_unit_to_usd_t(total, fx, [(1e6, "百万韩元"), (1e9, "十亿韩元"),
                                                        (1e8, "亿韩元"), (1e12, "兆韩元")], lo, hi, diag)
                d = f"{trd[:4]}-{trd[4:6]}-{trd[6:]}"
                return {"cap_usd_t": round(cap_t, 2), "index_anchor": kospi, "anchor_date": d,
                        "anchor_source": "KRX数据门户(官方快照)"}
        diag.append(f"KRX {trd}: 无完整数据，回退前一日")
    raise RuntimeError("KRX 连续7天无有效快照（可能美国IP不可达或字段变更）")


def anchor_us(cur, diag):
    """美股不改绝对锚（无免费自动源），只把浮动指数升级为 Wilshire 5000 全市场指数（若 Yahoo 有）。
    全市场指数自带新股/退市调整，漂移远小于标普500。"""
    if cur.get("index") in ("^FTW5000", "^W5000"):
        d, _ = yahoo_last(cur["index"])   # 健康检查：数据断供则报错（由人/AI 处理，不自动回退）
        if (datetime.date.today() - d).days > 10:
            raise RuntimeError(f"{cur['index']} 数据已停更（最后 {d}）")
        diag.append(f"{cur['index']} 健康（最后 {d}），无需变更")
        return None
    a_date = datetime.date.fromisoformat(cur.get("anchor_date") or "2026-07-09")
    for sym in ("^FTW5000", "^W5000"):
        try:
            last_d, _ = yahoo_last(sym)
            if (datetime.date.today() - last_d).days > 10:
                diag.append(f"{sym} 存在但已停更（最后 {last_d}），跳过")
                continue
            wd, wlvl = yahoo_close_on(sym, a_date)
            diag.append(f"迁移浮动指数 ^GSPC → {sym}：锚点日 {wd} 指数 {wlvl}（绝对锚 {cur['cap_usd_t']}T 不变）")
            return {"index": sym, "index_anchor": wlvl,
                    "anchor_source": f"绝对锚沿用人工校准，漂移指数已升级为{sym}"}
        except Exception as e:
            diag.append(f"{sym} 探测失败: {e}")
    diag.append("Yahoo 无可用 Wilshire 全市场指数，维持 ^GSPC 浮动")
    return None


# ---------------- 主流程 ----------------

def main():
    cfg = json.loads(CFG_PATH.read_text(encoding="utf-8"))
    stocks = {s["key"]: s for s in cfg["stocks"]}
    log = {"run": datetime.datetime.now(datetime.timezone.utc)
                  .astimezone(datetime.timezone(datetime.timedelta(hours=8)))
                  .strftime("%Y-%m-%d %H:%M 北京时间"), "markets": {}}
    changed = False
    # 韩国不在自动列表：KRX 数据门户 2025 改版后要求登录账号（见 pykrx auth.py），
    # 免费无鉴权路线已断；韩国绝对锚由桌面端 AI 定时任务月检（KOSPI 指数浮动仍全自动）。
    jobs = [("港股", anchor_hk), ("日本", anchor_jp), ("A股", anchor_cn)]
    for key, fn in jobs:
        diag, entry = [], {"ok": False}
        try:
            res = fn(diag)
            lo, hi = SANITY[key]
            if not (lo <= res["cap_usd_t"] <= hi):
                raise RuntimeError(f"出界：{res['cap_usd_t']}T ∉ [{lo},{hi}]")
            s = stocks[key]
            if (abs(s["cap_usd_t"] - res["cap_usd_t"]) > 0.005 or
                    abs(s["index_anchor"] - res["index_anchor"]) > 1e-6):
                s.update(res)
                changed = True
            else:
                s.update({k: v for k, v in res.items() if k in ("anchor_date", "anchor_source")})
            entry = {"ok": True, **res}
        except Exception as e:
            entry = {"ok": False, "error": str(e)[:400]}
            print(f"::warning::锚点校准失败 {key}: {e}", file=sys.stderr)
        entry["diag"] = diag
        log["markets"][key] = entry

    log["markets"]["韩国"] = {"ok": None, "skipped": "KRX数据门户已需登录（2025改版），"
                              "绝对锚由AI定时任务月检；KOSPI指数浮动不受影响"}

    # 美股：漂移指数升级探测
    diag, entry = [], {}
    try:
        res = anchor_us(stocks["美股"], diag)
        if res:
            stocks["美股"].update(res)
            changed = True
        entry = {"ok": True, "changed": bool(res)}
    except Exception as e:
        entry = {"ok": False, "error": str(e)[:400]}
        print(f"::warning::美股指数探测失败: {e}", file=sys.stderr)
    entry["diag"] = diag
    log["markets"]["美股"] = entry

    # 金融史页签的百年走势图数据：标普500全历史月线（1927至今）。
    # 坑（实测2026-07-12）：range=max&interval=1mo 会被 Yahoo 悄悄降级成 1984 年起的 168 个稀疏点
    # （与手册记载的 range=max 日线降级同源）。解法：显式 period1/period2 按 20 年分块抓，合并去重。
    # 防呆：新抓的点数不如现存文件就不覆盖（历史只会变多不会变少）。失败不影响锚点主流程。
    try:
        import time as _t
        bym = {}
        t0, now_ts = -1362000000, int(_t.time())          # 1926-11 起
        step = 20 * 365 * 86400
        while t0 < now_ts:
            t1 = min(t0 + step, now_ts)
            url = (f"https://query1.finance.yahoo.com/v8/finance/chart/%5EGSPC"
                   f"?period1={t0}&period2={t1}&interval=1mo")
            try:
                res = http_json(url)["chart"]["result"][0]
                for t, c in zip(res.get("timestamp") or [],
                                (res["indicators"]["quote"][0].get("close") or [])):
                    if c:
                        ym = datetime.datetime.fromtimestamp(t, datetime.timezone.utc).strftime("%Y-%m")
                        bym[ym] = round(float(c), 2)
            except Exception as e:
                print(f"  标普月线分块 {t0} 失败: {e}", file=sys.stderr)
            t0 = t1 + 86400
            _t.sleep(1)
        pts = sorted(bym.items())
        old_n = 0
        try:
            old_n = json.loads((ROOT / "data/spx_history.json").read_text(encoding="utf-8")).get("n", 0)
        except Exception:
            pass
        if len(pts) >= max(old_n, 500):
            (ROOT / "data/spx_history.json").write_text(json.dumps(
                {"sym": "^GSPC", "updated": log["run"], "n": len(pts), "points": pts},
                ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
            print(f"标普百年月线：{len(pts)} 个月（{pts[0][0]} ~ {pts[-1][0]}）")
        else:
            print(f"::warning::标普月线仅 {len(pts)} 点（现存 {old_n}），疑似降级，不覆盖")
    except Exception as e:
        print(f"::warning::标普百年月线生成失败: {e}", file=sys.stderr)

    if changed:
        dates = [s.get("anchor_date") for s in cfg["stocks"] if s.get("anchor_date")]
        cfg["anchor_date"] = max(dates) if dates else cfg.get("anchor_date", "")
        cfg["_口径"] = ("股市＝官方总市值锚点 × 各自大盘指数相对锚点日的涨跌（锚点每周由 update_anchors.py "
                      "自动校准：港股HKMA/日本JPX/A股沪深交易所/韩国KRX 官方数据；美股绝对锚人工+AI月检）。"
                      "金银＝现货价 × 地上存量（黄金216,265吨/WGC，白银1,751,000吨/CPM）。比特币＝CoinGecko 实时市值。")
        CFG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    LOG_PATH.write_text(json.dumps(log, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    ok = sum(1 for m in log["markets"].values() if m.get("ok"))
    print(f"校锚完成：{ok}/{len(log['markets'])} 成功，config {'已更新' if changed else '无变化'}")
    for k, m in log["markets"].items():
        print(f"  {k}: {'✓' if m.get('ok') else '✗ ' + m.get('error', '')}")


if __name__ == "__main__":
    main()

# 触发校锚工作流验证ECS路线
